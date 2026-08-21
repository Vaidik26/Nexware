import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from io import BytesIO
from typing import Any, List, Dict
from datetime import datetime


def parse_catalogue_excel(file_content: bytes) -> list[dict]:
    try:
        df = pd.read_excel(BytesIO(file_content))
    except Exception as e:
        raise ValueError(f"Invalid Excel document file structure or formatting: {str(e)}")
        
    if df.empty:
        raise ValueError("Uploaded Excel sheet is completely empty.")
        
    # Normalize columns to handle minor header variation
    col_map = {col.lower().strip(): col for col in df.columns}
    item_num_col = col_map.get("item number", col_map.get("code", "Item Number"))
    item_name_col = col_map.get("item name", col_map.get("description", "Item Name"))
    barcode_col = col_map.get("barcode", "Barcode")
    unit_col = col_map.get("unit", col_map.get("uom", "Unit"))

    items = []
    for _, row in df.iterrows():
        item_no = str(row.get(item_num_col, "")).strip() if item_num_col in row else ""
        barcode = str(row.get(barcode_col, "")).strip() if barcode_col in row else ""
        if item_no == "nan": item_no = ""
        if barcode == "nan": barcode = ""
        if item_no and barcode:
            items.append({
                "item_number": item_no,
                "item_name": str(row.get(item_name_col, "Unnamed SKU")).strip() if item_name_col in row else "Unnamed SKU",
                "barcode": barcode,
                "unit": str(row.get(unit_col, "PCS")).strip() if unit_col in row else "PCS",
            })
    if not items:
        raise ValueError("No valid rows found in Excel. Make sure 'Item Number' and 'Barcode' columns are present and filled.")
    return items


def generate_catalogue_excel(items: list) -> bytes:
    data = [
        {
            "Item Number": i.item_number,
            "Item Name": i.item_name,
            "Barcode": i.barcode,
            "Unit": i.unit,
        }
        for i in items
    ]
    df = pd.DataFrame(data)
    buf = BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Sales Items")
    return buf.getvalue()


def generate_branded_picklist_excel(items_list: Any) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Warehouse Pick List"
    
    # Enable display grid lines explicitly
    ws.views.sheetView[0].showGridLines = True
    
    # Enterprise Brand Colors
    brand_green = "154C34"       # Deep emerald brand header banner
    header_green = "2B4C3A"      # Sage green table header
    zebra_light = "F4F7F5"       # Alternating row background fill
    border_gray = "CCCCCC"       # Crisp cell borders
    white = "FFFFFF"
    
    thin_side = Side(border_style="thin", color=border_gray)
    cell_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    
    # Row 1: Executive Brand Header Banner (Merged A1:F1)
    ws.merge_cells("A1:F1")
    title_cell = ws["A1"]
    title_cell.value = "NEXWARE ENTERPRISE OS — WAREHOUSE FLOOR PICK LIST"
    title_cell.font = Font(name="Arial", size=15, bold=True, color=white)
    title_cell.fill = PatternFill(start_color=brand_green, end_color=brand_green, fill_type="solid")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 36
    
    # Row 2: Date Generated Tag (Merged A2:F2)
    ws.merge_cells("A2:F2")
    customer_name = getattr(items_list, 'customer_name', '') if not isinstance(items_list, (dict, list)) else (items_list.get('customer_name', '') if isinstance(items_list, dict) else '')
    date_cell = ws["A2"]
    date_cell.value = f"Customer: {customer_name or 'N/A'} | Date Generated: {datetime.now().strftime('%B %d, %Y')}"
    date_cell.font = Font(name="Arial", size=11, bold=True, italic=True, color="333333")
    date_cell.fill = PatternFill(start_color="E8F2EC", end_color="E8F2EC", fill_type="solid")
    date_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[2].height = 24
    
    # Row 3: Spacer
    ws.row_dimensions[3].height = 10
    
    # Row 4: Operational Table Headers
    headers = ["SI", "Bin Location", "Packaging", "Item Code (Barcode)", "Description / Product Title", "Quantity", "Checked", "Picked"]
    for col_idx, h_text in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col_idx, value=h_text)
        cell.font = Font(name="Arial", size=11, bold=True, color=white)
        cell.fill = PatternFill(start_color=header_green, end_color=header_green, fill_type="solid")
        cell.alignment = Alignment(horizontal="center" if col_idx != 3 else "left", vertical="center", indent=1 if col_idx == 3 else 0)
        cell.border = cell_border
    ws.row_dimensions[4].height = 28
    
    pl_status = getattr(items_list, 'status', '') if not isinstance(items_list, (dict, list)) else (items_list.get('status', '') if isinstance(items_list, dict) else '')

    # Handle input whether it's an ORM object from DB or a memory dictionary from frontend upload
    if hasattr(items_list, 'items'):
        raw = items_list.items
        items_data = [
            {
                "barcode": getattr(i, "barcode", ""),
                "product_name": getattr(i, "product_name", ""),
                "quantity": getattr(i, "quantity", 1) or 1,
                "is_picked": getattr(i, "is_picked", False),
                "is_full_carton": getattr(i, "is_full_carton", False),
                "bin_location": getattr(i, "bin_location", "") or "",
            }
            for i in raw
        ]
    elif isinstance(items_list, list):
        items_data = [
            {
                "barcode": str(i.get("barcode", "") or i.get("item_number", "")).strip(),
                "product_name": str(i.get("product_name", "") or i.get("itemName", "") or i.get("description", "")).strip(),
                "quantity": float(i.get("quantity", 1) or 1),
                "is_picked": i.get("is_picked", False),
                "is_full_carton": i.get("is_full_carton", False),
                "bin_location": str(i.get("bin_location", "")).strip(),
            }
            for i in items_list
        ]
    else:
        items_data = []
        
    if not items_data:
        raise ValueError("Cannot generate operational Excel picklist: Order has no items attached.")
        
    # Data Rows (Row 5+) with Zebra Striping and strict string barcode formula formatting
    for idx, item in enumerate(items_data, start=1):
        row_num = 4 + idx
        bg_color = white if idx % 2 != 0 else zebra_light
        row_fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
        
        # SI
        c1 = ws.cell(row=row_num, column=1, value=idx)
        c1.alignment = Alignment(horizontal="center", vertical="center")
        
        # Bin Location
        bin_loc = str(item.get("bin_location", "")) or "N/A"
        c2 = ws.cell(row=row_num, column=2, value=bin_loc.upper())
        c2.alignment = Alignment(horizontal="center", vertical="center")
        c2.font = Font(name="Arial", size=11, bold=True, color="1D5D42")

        # Packaging (Full Carton / Loose Item)
        pkg_type = "Full Carton" if item.get("is_full_carton") else "Loose Item"
        c3 = ws.cell(row=row_num, column=3, value=pkg_type)
        c3.alignment = Alignment(horizontal="center", vertical="center")
        c3.font = Font(name="Arial", size=10, bold=True, color="000000" if item.get("is_full_carton") else "555555")
        
        # Barcode (Strict text format '@' totally prevents scientific notation 6.29E+12)
        bc_str = str(item.get("barcode", "")).strip() or "N/A"
        c4 = ws.cell(row=row_num, column=4)
        c4.number_format = '@'
        c4.data_type = 's'
        c4.value = bc_str
        c4.font = Font(name="Consolas", size=11, bold=True)
        c4.alignment = Alignment(horizontal="center", vertical="center")
        
        # Description (with text wrapping enabled for long item titles)
        prod_title = str(item.get("product_name", ""))
        c5 = ws.cell(row=row_num, column=5, value=prod_title)
        c5.alignment = Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True)
        c5.font = Font(name="Arial", size=11, bold=True)
        
        # Quantity (Default to 1 if missing or 0)
        qty_val = float(item.get("quantity", 1)) if item.get("quantity") else 1.0
        c6 = ws.cell(row=row_num, column=6, value=qty_val if qty_val % 1 != 0 else int(qty_val))
        c6.alignment = Alignment(horizontal="center", vertical="center")
        c6.font = Font(name="Arial", size=12, bold=True)
        
        is_picked_row = pl_status in ("waiting_verification", "picked", "verified", "completed") or item.get("is_picked", False)
        is_checked_row = pl_status in ("verified", "completed")

        c7_val = "[ ✔ ]" if is_checked_row else "[        ]"
        c8_val = "[ ✔ ]" if is_picked_row else "[        ]"

        # Checked & Picked slots
        c7 = ws.cell(row=row_num, column=7, value=c7_val)
        c7.alignment = Alignment(horizontal="center", vertical="center")
        c7.font = Font(name="Arial" if is_checked_row else "Consolas", size=11, bold=is_checked_row, color="154c34" if is_checked_row else "555555")
        
        c8 = ws.cell(row=row_num, column=8, value=c8_val)
        c8.alignment = Alignment(horizontal="center", vertical="center")
        c8.font = Font(name="Arial" if is_picked_row else "Consolas", size=11, bold=is_picked_row, color="154c34" if is_picked_row else "555555")
        
        for cell in [c1, c2, c3, c4, c5, c6, c7, c8]:
            cell.fill = row_fill
            cell.border = cell_border
            
        # Dynamically adjust row height if item title is long and wraps onto multiple lines
        lines_needed = max(1, (len(prod_title) // 50) + 1)
        ws.row_dimensions[row_num].height = max(26, lines_needed * 20)
        
    # Pre-Adjusted Automatic Column Widths (Zero manual dragging or text clipping!)
    max_desc_len = max([len(str(item.get("product_name", ""))) for item in items_data], default=30)
    col_c_width = min(95, max(68, int(max_desc_len * 0.95)))
    
    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = col_c_width
    ws.column_dimensions["F"].width = 12
    ws.column_dimensions["G"].width = 14
    ws.column_dimensions["H"].width = 14
    
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def generate_picklist_excel(picklist: Any) -> bytes:
    return generate_branded_picklist_excel(picklist)


def generate_branded_price_history_excel(payload: dict) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Market Price Intelligence"
    ws.views.sheetView[0].showGridLines = True

    # Enterprise Brand Colors
    brand_green = "154C34"       # Deep emerald brand header banner
    date_green = "1D5D42"        # Medium emerald for daily table date section titles
    header_green = "2B4C3A"      # Sage green table column headers
    zebra_light = "F4F7F5"       # Alternating row background fill
    border_gray = "CCCCCC"       # Crisp cell borders
    white = "FFFFFF"

    thin_side = Side(border_style="thin", color=border_gray)
    cell_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    # Row 1: Executive Brand Header Banner (Merged A1:F1)
    ws.merge_cells("A1:F1")
    title_cell = ws["A1"]
    title_cell.value = "NEXWARE ENTERPRISE OS — COMMODITY PRICE INTELLIGENCE REPORT"
    title_cell.font = Font(name="Arial", size=15, bold=True, color=white)
    title_cell.fill = PatternFill(start_color=brand_green, end_color=brand_green, fill_type="solid")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 36

    # Row 2: Scope & Timestamp Tag (Merged A2:F2)
    ws.merge_cells("A2:F2")
    date_cell = ws["A2"]
    scope_str = str(payload.get("scope", "ALL")).upper()
    date_cell.value = f"Scope / Time Range: {scope_str} | Generated On: {datetime.now().strftime('%B %d, %Y - %I:%M %p')}"
    date_cell.font = Font(name="Arial", size=11, bold=True, italic=True, color="333333")
    date_cell.fill = PatternFill(start_color="E8F2EC", end_color="E8F2EC", fill_type="solid")
    date_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[2].height = 24

    current_row = 3
    dates_list = payload.get("dates", [])
    if not dates_list:
        ws.cell(row=current_row + 1, column=1, value="No historical pricing records found for this scope.")
        buf = BytesIO()
        wb.save(buf)
        return buf.getvalue()

    for date_idx, date_block in enumerate(dates_list):
        if date_idx > 0:
            ws.row_dimensions[current_row].height = 12
            current_row += 1

        # Date Section Title Banner
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=6)
        d_cell = ws.cell(row=current_row, column=1)
        d_str = date_block.get("date", "")
        d_fmt = date_block.get("date_formatted", d_str)
        d_cell.value = f"📅 MARKET RECORD DATE: {d_fmt} ({d_str})"
        d_cell.font = Font(name="Arial", size=12, bold=True, color=white)
        d_cell.fill = PatternFill(start_color=date_green, end_color=date_green, fill_type="solid")
        d_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[current_row].height = 28
        current_row += 1

        # Column Headers
        headers = ["S.No", "Commodity Item Name", "Bag/Carton Weight", "Local Dubai Price", "International CIF ($)", "International FOB ($)"]
        for col_idx, h_text in enumerate(headers, start=1):
            cell = ws.cell(row=current_row, column=col_idx, value=h_text)
            cell.font = Font(name="Arial", size=11, bold=True, color=white)
            cell.fill = PatternFill(start_color=header_green, end_color=header_green, fill_type="solid")
            cell.alignment = Alignment(
                horizontal="center" if col_idx in [1, 3] else ("left" if col_idx == 2 else "right"),
                vertical="center",
                indent=1 if col_idx == 2 else 0
            )
            cell.border = cell_border
        ws.row_dimensions[current_row].height = 26
        current_row += 1

        # Data Rows
        rows_data = date_block.get("rows", [])
        for idx, row_item in enumerate(rows_data, start=1):
            bg_color = white if idx % 2 != 0 else zebra_light
            row_fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")

            c1 = ws.cell(row=current_row, column=1, value=row_item.get("sno", idx))
            c1.alignment = Alignment(horizontal="center", vertical="center")
            c1.font = Font(name="Arial", size=11)

            c2 = ws.cell(row=current_row, column=2, value=row_item.get("commodity", ""))
            c2.alignment = Alignment(horizontal="left", vertical="center", indent=1)
            c2.font = Font(name="Arial", size=11, bold=True)

            c3 = ws.cell(row=current_row, column=3, value=row_item.get("weight", "N/A"))
            c3.alignment = Alignment(horizontal="center", vertical="center")
            c3.font = Font(name="Consolas", size=11, bold=True, color="444444")

            c4 = ws.cell(row=current_row, column=4, value=row_item.get("local_price", "—"))
            c4.alignment = Alignment(horizontal="right", vertical="center", indent=1)
            c4.font = Font(name="Arial", size=11, bold=True, color="005530" if row_item.get("local_price", "—") != "—" else "888888")

            c5 = ws.cell(row=current_row, column=5, value=row_item.get("cif_price", "—"))
            c5.alignment = Alignment(horizontal="right", vertical="center", indent=1)
            c5.font = Font(name="Arial", size=11, bold=True)

            c6 = ws.cell(row=current_row, column=6, value=row_item.get("fob_price", "—"))
            c6.alignment = Alignment(horizontal="right", vertical="center", indent=1)
            c6.font = Font(name="Arial", size=11, bold=True)

            for cell in [c1, c2, c3, c4, c5, c6]:
                cell.fill = row_fill
                cell.border = cell_border

            ws.row_dimensions[current_row].height = 24
            current_row += 1

        current_row += 1 # Spacer after daily table

    # Auto-Adjusted Column Widths with generous padding
    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 44
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 24
    ws.column_dimensions["E"].width = 24
    ws.column_dimensions["F"].width = 24

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()



from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Protection

def generate_price_capture_template(materials: list, market_type: str) -> bytes:
    wb = openpyxl.Workbook()
    
    # Define styles (Nexware Green Theme)
    header_fill = PatternFill(start_color="064E3B", end_color="064E3B", fill_type="solid") # Dark Green
    header_font = Font(color="FFFFFF", bold=True, size=11)
    
    row_fill_alt = PatternFill(start_color="ECFDF5", end_color="ECFDF5", fill_type="solid") # Light Emerald
    row_fill_normal = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    
    thin_border = Border(
        left=Side(style='thin', color="CCCCCC"), 
        right=Side(style='thin', color="CCCCCC"), 
        top=Side(style='thin', color="CCCCCC"), 
        bottom=Side(style='thin', color="CCCCCC")
    )
    
    locked_style = Protection(locked=True)
    unlocked_style = Protection(locked=False)
    
    def create_sheet(sheet_name, market_filter, ws_index):
        if ws_index == 0:
            ws = wb.active
            ws.title = sheet_name
        else:
            ws = wb.create_sheet(title=sheet_name)
            
        # Determine columns
        cols = ["S.No", "SKU / Index Code", "Commodity Item Name", "Category", "Bag/CTN Weight"]
        
        if market_filter == "DXB":
            cols.extend(["Local Dubai Price (AED)", "Supplier (Dubai)", "Local Oman Price (OMR)", "Supplier (Oman)"])
        elif market_filter == "INT":
            cols.extend(["International CIF (USD)", "International FOB (USD)", "Supplier (INT)"])
        else: # BOTH
            cols.extend(["Local Dubai Price (AED)", "Supplier (Dubai)", "Local Oman Price (OMR)", "Supplier (Oman)", "International CIF (USD)", "International FOB (USD)", "Supplier (INT)"])
            
        # Write Headers
        ws.append(cols)
        for col_idx, cell in enumerate(ws[1], 1):
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border
            
        # Filter materials
        sheet_mats = [m for m in materials if m.market_type == market_filter]
        
        for idx, mat in enumerate(sheet_mats, 1):
            row_data = [
                idx,
                mat.material_code,
                mat.material_name,
                mat.category,
                mat.bag_carton_weight if mat.bag_carton_weight else "-",
            ]
                
            # Prices (blank for entry)
            if market_filter == "DXB":
                row_data.extend(["", "", "", ""])
            elif market_filter == "INT":
                row_data.extend(["", ""])
            else:
                row_data.extend(["", "", "", "", "", ""])
                
            ws.append(row_data)
            current_row = ws[ws.max_row]
            
            fill = row_fill_alt if idx % 2 == 0 else row_fill_normal
            
            for c_idx, cell in enumerate(current_row, 1):
                cell.fill = fill
                cell.border = thin_border
                cell.alignment = Alignment(vertical="center")
                
                # Lock columns 1 to 5
                if c_idx <= 5:
                    cell.protection = locked_style
                else:
                    cell.protection = unlocked_style
                
        ws.protection.sheet = True
        ws.protection.password = ""
        
        # Adjust column widths
        ws.column_dimensions["A"].width = 8
        ws.column_dimensions["B"].width = 18
        ws.column_dimensions["C"].width = 40
        ws.column_dimensions["D"].width = 20
        ws.column_dimensions["E"].width = 18
        if market_filter == "DXB":
            ws.column_dimensions["F"].width = 22
            ws.column_dimensions["G"].width = 20
            ws.column_dimensions["H"].width = 22
            ws.column_dimensions["I"].width = 20
        elif market_filter == "INT":
            ws.column_dimensions["F"].width = 22
            ws.column_dimensions["G"].width = 22
        else:
            ws.column_dimensions["F"].width = 22
            ws.column_dimensions["G"].width = 20
            ws.column_dimensions["H"].width = 22
            ws.column_dimensions["I"].width = 20
            ws.column_dimensions["J"].width = 22
            ws.column_dimensions["K"].width = 22
            
    if market_type == "ALL":
        create_sheet("Dubai Local", "DXB", 0)
        create_sheet("International", "INT", 1)
        create_sheet("Both Markets", "BOTH", 2)
    elif market_type == "DXB":
        create_sheet("Dubai Local", "DXB", 0)
    elif market_type == "INT":
        create_sheet("International", "INT", 0)
    elif market_type == "BOTH":
        create_sheet("Both Markets", "BOTH", 0)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
